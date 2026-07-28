"""Extracción de movimientos desde archivos Excel (.xlsx / .xlsm) y CSV.

Los bancos peruanos exportan el estado de cuenta a Excel con encabezados en español,
pero no siempre en la primera fila (suele haber un bloque de datos del titular arriba).
Por eso se busca la fila de encabezados en las primeras filas de cada hoja.
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date, datetime
from decimal import Decimal

from openpyxl import load_workbook

from .pdf_extractor import (
    CABECERAS,
    ErrorExtraccion,
    NoReconocidoError,
    ResultadoExtraccion,
    _es_ruido,
    _normalizar,
    _signo_por_texto,
    parsear_fecha,
    parsear_monto,
)

logger = logging.getLogger(__name__)

MAX_FILAS_BUSCAR_CABECERA = 25


def _clasificar_cabecera(texto: str) -> str | None:
    """Mapea el texto de una celda de encabezado a un tipo de columna conocido."""
    n = _normalizar(str(texto)).replace(".", "").replace(":", "").replace(" ", "")
    if not n:
        return None
    for tipo, alias in CABECERAS.items():
        if n in alias:
            return tipo
    # Coincidencia parcial: "fechadeoperacion", "montodelcargo", ...
    for tipo, alias in CABECERAS.items():
        if any(a in n for a in alias if len(a) > 4):
            return tipo
    return None


def _mapear_columnas(fila: list) -> dict[str, int] | None:
    """Devuelve {tipo_columna: índice} si la fila parece un encabezado."""
    mapa: dict[str, int] = {}
    for indice, celda in enumerate(fila):
        if celda is None:
            continue
        tipo = _clasificar_cabecera(celda)
        if tipo and tipo not in mapa:
            mapa[tipo] = indice

    if "fecha" in mapa and ({"cargo", "abono", "monto"} & mapa.keys()):
        return mapa
    return None


def _valor_celda_fecha(valor, anio_defecto: int) -> date | None:
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    if valor is None:
        return None
    return parsear_fecha(str(valor).strip(), anio_defecto)


def _valor_celda_monto(valor) -> Decimal | None:
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    return parsear_monto(str(valor))


def _filas_a_movimientos(
    filas: list[list], mapa: dict[str, int], anio_defecto: int
) -> list[dict]:
    movimientos: list[dict] = []

    for fila in filas:
        def celda(tipo: str):
            indice = mapa.get(tipo)
            if indice is None or indice >= len(fila):
                return None
            return fila[indice]

        fecha = _valor_celda_fecha(celda("fecha"), anio_defecto)
        if fecha is None:
            continue

        descripcion = str(celda("descripcion") or "").strip()
        if _es_ruido(descripcion) or _es_ruido(" ".join(str(c) for c in fila if c)):
            continue

        cargo = _valor_celda_monto(celda("cargo"))
        abono = _valor_celda_monto(celda("abono"))
        monto_col = _valor_celda_monto(celda("monto"))

        if cargo and abono:
            monto = abono - abs(cargo)
        elif cargo:
            monto = -abs(cargo)
        elif abono:
            monto = abs(abono)
        elif monto_col is not None:
            monto = monto_col
            if monto > 0 and _signo_por_texto(descripcion) < 0:
                monto = -monto
        else:
            continue

        if monto == 0:
            continue

        movimientos.append(
            {
                "fecha": fecha,
                "monto": Decimal(str(monto)).quantize(Decimal("0.01")),
                "descripcion": descripcion or "Movimiento sin descripción",
            }
        )

    return movimientos


def _extraer_de_hojas(hojas: list[tuple[str, list[list]]], anio_defecto: int) -> list[dict]:
    movimientos: list[dict] = []
    for nombre, filas in hojas:
        mapa = None
        inicio = 0
        for indice, fila in enumerate(filas[:MAX_FILAS_BUSCAR_CABECERA]):
            mapa = _mapear_columnas(fila)
            if mapa:
                inicio = indice + 1
                break
        if not mapa:
            logger.debug("Hoja '%s' sin encabezado reconocible, se omite", nombre)
            continue
        movimientos.extend(_filas_a_movimientos(filas[inicio:], mapa, anio_defecto))
    return movimientos


def extraer_movimientos_excel(contenido: bytes, nombre_archivo: str = "") -> ResultadoExtraccion:
    """Lee un Excel o CSV de estado de cuenta y devuelve sus movimientos.

    Returns:
        ResultadoExtraccion con `movimientos` = [{fecha, monto, descripcion}].

    Raises:
        NoReconocidoError: no se encontró una tabla de movimientos.
    """
    resultado = ResultadoExtraccion()
    anio_defecto = date.today().year

    if nombre_archivo.lower().endswith(".csv"):
        try:
            texto = contenido.decode("utf-8-sig")
        except UnicodeDecodeError:
            texto = contenido.decode("latin-1")
        # Los bancos peruanos exportan CSV con ; o con ,
        delimitador = ";" if texto.count(";") > texto.count(",") else ","
        filas = [list(f) for f in csv.reader(io.StringIO(texto), delimiter=delimitador)]
        hojas = [("csv", filas)]
    else:
        try:
            libro = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
        except Exception as exc:
            raise ErrorExtraccion(
                f"El archivo no se pudo abrir como Excel: {exc}", "EXCEL_INVALIDO"
            ) from exc
        hojas = [
            (hoja.title, [list(fila) for fila in hoja.iter_rows(values_only=True)])
            for hoja in libro.worksheets
        ]
        libro.close()

    movimientos = _extraer_de_hojas(hojas, anio_defecto)

    vistos: set[tuple] = set()
    for movimiento in movimientos:
        clave = (movimiento["fecha"], movimiento["monto"], movimiento["descripcion"])
        if clave in vistos:
            continue
        vistos.add(clave)
        resultado.movimientos.append(movimiento)

    if not resultado.movimientos:
        raise NoReconocidoError(
            "No se encontraron movimientos en el archivo. Debe tener columnas de "
            "fecha, descripción y monto (o cargo/abono)."
        )

    resultado.movimientos.sort(key=lambda m: m["fecha"])
    logger.info("Excel procesado: movimientos=%s", len(resultado.movimientos))
    return resultado
